
import os
import glob
import pandas as pd

def test_excel_read():
    # Find xlsx files in the current directory
    xlsx_files = glob.glob('*.xlsx')
    
    if not xlsx_files:
        print("No .xlsx files found in the current directory.")
        print("Please place the Excel file you want to test in this folder:", os.getcwd())
        return

    # Use the first found file, or prefer one with '분개장' in the name if multiple exist
    target_file = xlsx_files[0]
    for f in xlsx_files:
        if '분개장' in f:
            target_file = f
            break
            
    print(f"--- Reading file: {target_file} ---")
    
    try:
        # Load the excel file
        # Using pandas is easiest for "top 10 lines"
        # We need to install openpyxl if not present, but let's assume it is or pandas handles it
        xl = pd.ExcelFile(target_file)
        
        print(f"Sheet names: {xl.sheet_names}")
        
        if len(xl.sheet_names) > 0:
            first_sheet = xl.sheet_names[0]
            print(f"\n--- Reading top 10 rows from sheet: {first_sheet} ---")
            
            df = pd.read_excel(target_file, sheet_name=first_sheet, nrows=10, header=None)
            print(df.to_string())
            
    except Exception as e:
        print(f"Error reading file: {e}")
        try:
            import openpyxl
            print("\nAttempting with openpyxl directly...")
            wb = openpyxl.load_workbook(target_file)
            print(f"Sheet names: {wb.sheetnames}")
            ws = wb.active
            print(f"\n--- Reading top 10 rows from sheet: {ws.title} ---")
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 10: break
                print(row)
        except ImportError:
            print("openpyxl is not installed. Please run: pip install pandas openpyxl")
        except Exception as e2:
             print(f"Error with openpyxl: {e2}")

if __name__ == "__main__":
    test_excel_read()
